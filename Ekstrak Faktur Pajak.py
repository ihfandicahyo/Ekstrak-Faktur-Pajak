import pdfplumber
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def clean_val(val):
    if val is None:
        return "-"
    s_val = str(val).strip()
    return s_val if s_val else "-"

def parse_currency(value):
    if not value:
        return 0
    clean = re.sub(r'[^\d,]', '', str(value))
    if not clean:
        return 0
    return clean

def auto_adjust_excel_width(filename):
    wb = load_workbook(filename)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                val_len = len(str(cell.value))
                if val_len > max_length:
                    max_length = val_len
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50: adjusted_width = 50
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    wb.save(filename)

def extract_faktur_data(pdf_path):
    data_list = []
    header_info = {}
    
    with pdfplumber.open(pdf_path) as pdf:
        full_text_all_pages = ""
        for page in pdf.pages:
            full_text_all_pages += page.extract_text() + "\n"
        
        page_one_text = pdf.pages[0].extract_text()
        
        match_faktur = re.search(r"Kode dan Nomor Seri Faktur Pajak:\s*([\d]+)", page_one_text)
        header_info['Nomor Faktur'] = clean_val(match_faktur.group(1)) if match_faktur else "-"
        
        match_ref = re.search(r"Referensi\s*:?\s*([0-9a-zA-Z]+)", full_text_all_pages)
        header_info['Referensi'] = clean_val(match_ref.group(1)) if match_ref else "-"
        
        idx_pkp = page_one_text.find("Pengusaha Kena Pajak")
        idx_pembeli = page_one_text.find("Pembeli Barang Kena Pajak")
        
        idx_table = -1
        for keyword in ["The following table", "Kode Barang", "Nama Barang"]:
            pos = page_one_text.find(keyword, idx_pembeli)
            if pos != -1:
                idx_table = pos
                break
        if idx_table == -1: idx_table = len(page_one_text)

        pkp_block = page_one_text[idx_pkp:idx_pembeli] if idx_pkp != -1 and idx_pembeli != -1 else ""
        pembeli_block = page_one_text[idx_pembeli:idx_table] if idx_pembeli != -1 else ""

        match_pkp_nama = re.search(r"Nama\s*:\s*(.*)", pkp_block)
        header_info['PKP Nama'] = clean_val(match_pkp_nama.group(1)) if match_pkp_nama else "-"
        
        match_pkp_alamat = re.search(r"Alamat\s*:\s*(.*?)(?=NPWP)", pkp_block, re.DOTALL)
        header_info['PKP Alamat'] = clean_val(match_pkp_alamat.group(1).replace('\n', ' ')) if match_pkp_alamat else "-"
        
        match_pkp_npwp = re.search(r"NPWP\s*:\s*([\d]+)", pkp_block)
        header_info['PKP NPWP'] = clean_val(match_pkp_npwp.group(1)) if match_pkp_npwp else "-"

        match_pembeli_nama = re.search(r"Nama\s*:\s*(.*)", pembeli_block)
        header_info['Pembeli Nama'] = clean_val(match_pembeli_nama.group(1)) if match_pembeli_nama else "-"
        
        match_pembeli_alamat = re.search(r"Alamat\s*:\s*(.*?)(?=NPWP)", pembeli_block, re.DOTALL)
        header_info['Pembeli Alamat'] = clean_val(match_pembeli_alamat.group(1).replace('\n', ' ')) if match_pembeli_alamat else "-"
        
        match_pembeli_npwp = re.search(r"NPWP\s*:\s*([\d]+)", pembeli_block)
        header_info['Pembeli NPWP'] = clean_val(match_pembeli_npwp.group(1)) if match_pembeli_npwp else "-"
        
        match_pembeli_nik = re.search(r"NIK\s*:\s*([\d\-]*)", pembeli_block)
        header_info['Pembeli NIK'] = clean_val(match_pembeli_nik.group(1)) if match_pembeli_nik else "-"
        
        match_paspor = re.search(r"Nomor Paspor\s*:\s*([\w\-]*)", pembeli_block)
        header_info['Nomor Paspor'] = clean_val(match_paspor.group(1)) if match_paspor else "-"
        
        match_identitas = re.search(r"Identitas Lain\s*:\s*([\w\-]*)", pembeli_block)
        header_info['Identitas Lain'] = clean_val(match_identitas.group(1)) if match_identitas else "-"
        
        match_email = re.search(r"Email\s*:\s*(.*)", pembeli_block)
        header_info['Email Pembeli'] = clean_val(match_email.group(1)) if match_email else "-"
        
        matches_jual = re.findall(r"Harga Jual / Penggantian / Uang Muka / Termin[^\d\n]*\n?.*?([\d\.,]+)", full_text_all_pages, re.DOTALL)
        if matches_jual:
            valid_jual = [m for m in matches_jual if ',' in m or '.' in m]
            header_info['Total Harga Jual'] = parse_currency(valid_jual[-1]) if valid_jual else "0"
        else:
            header_info['Total Harga Jual'] = "0"
        
        match_potongan = re.search(r"Dikurangi Potongan Harga\s*([\d\.,]+)", full_text_all_pages)
        header_info['Total Potongan'] = parse_currency(match_potongan.group(1)) if match_potongan else "0"
        
        match_dpp = re.search(r"Dasar Pengenaan Pajak\s*([\d\.,]+)", full_text_all_pages)
        header_info['DPP'] = parse_currency(match_dpp.group(1)) if match_dpp else "0"
        
        match_ppn = re.search(r"Jumlah PPN.*?Nilai\)\s*([\d\.,]+)", full_text_all_pages)
        header_info['PPN'] = parse_currency(match_ppn.group(1)) if match_ppn else "0"
        
        match_ppnbm_tot = re.search(r"Jumlah PPnBM.*?Mewah\)\s*([\d\.,]+)", full_text_all_pages)
        header_info['PPnBM Total'] = parse_currency(match_ppnbm_tot.group(1)) if match_ppnbm_tot else "0"
        
        match_loc_date = re.search(r"(KOTA [A-Z ]+),\s*(\d{2}\s+[A-Za-z]+\s+\d{4})", full_text_all_pages)
        header_info['Lokasi'] = clean_val(match_loc_date.group(1)) if match_loc_date else "-"
        header_info['Tanggal'] = clean_val(match_loc_date.group(2)) if match_loc_date else "-"
        
        match_signer = re.search(r"Ditandatangani secara elektronik\s*\n\s*(.*)", full_text_all_pages)
        header_info['Penandatangan'] = clean_val(match_signer.group(1)) if match_signer else "-"

        all_tables = []
        for p in pdf.pages:
            t = p.extract_table()
            if t: all_tables.extend(t)

        for row in all_tables:
            if not row or len(row) < 3: continue
            
            row_str = "".join([str(x) for x in row])
            if "Nama Barang" in row_str or "Harga Jual" in row_str: continue
            
            raw_desc = str(row[2]) if len(row) > 2 else ""
            if not re.search(r"Rp\s*[\d\.,]+\s*x", raw_desc):
                continue

            raw_kode = clean_val(row[1]) if len(row) > 1 else "-"
            raw_total_str = str(row[-1]).strip() if len(row) > 0 else "0"

            item_pattern = re.compile(r"(.*?)\n?Rp\s*([\d\.,]+)\s*x\s*([\d\.,]+)\s*([A-Za-z]+).*?(?:Potongan Harga\s*=\s*Rp\s*([\d\.,]+))?.*?PPnBM\s*\(([\d\.,]+)%\)", re.DOTALL)
            
            items_found = list(item_pattern.finditer(raw_desc))
            
            if len(items_found) > 1:
                totals_in_col = str(row[-1]).split('\n')
                totals_in_col = [t for t in totals_in_col if len(re.sub(r'[^\d]', '', str(t))) > 0]
                
                for i, match in enumerate(items_found):
                    item_data = header_info.copy()
                    item_data['Kode Barang'] = raw_kode
                    item_data['Nama Barang'] = clean_val(match.group(1).replace('\n', ' '))
                    item_data['Harga Satuan'] = parse_currency(match.group(2))
                    item_data['Qty'] = parse_currency(match.group(3))
                    item_data['Satuan'] = clean_val(match.group(4))
                    item_data['Potongan Harga Item'] = parse_currency(match.group(5)) if match.group(5) else 0
                    item_data['PPnBM Persen'] = match.group(6) if match.group(6) else "0,00"
                    
                    try:
                        val_total = totals_in_col[i] if i < len(totals_in_col) else raw_total_str
                        item_data['Harga Total Item'] = parse_currency(val_total)
                    except:
                        item_data['Harga Total Item'] = 0
                    
                    data_list.append(item_data)
            
            else:
                item_data = header_info.copy()
                item_data['Kode Barang'] = raw_kode
                
                parts = raw_desc.split("Rp")
                item_data['Nama Barang'] = clean_val(parts[0].replace('\n', ' '))
                
                match_price_qty = re.search(r"Rp\s*([\d\.,]+)\s*x\s*([\d\.,]+)\s*([A-Za-z]+)", raw_desc)
                if match_price_qty:
                    item_data['Harga Satuan'] = parse_currency(match_price_qty.group(1))
                    item_data['Qty'] = parse_currency(match_price_qty.group(2))
                    item_data['Satuan'] = clean_val(match_price_qty.group(3))
                else:
                    item_data['Harga Satuan'] = 0
                    item_data['Qty'] = 0
                    item_data['Satuan'] = "-"
                
                match_disc = re.search(r"Potongan Harga\s*=\s*Rp\s*([\d\.,]+)", raw_desc)
                item_data['Potongan Harga Item'] = parse_currency(match_disc.group(1)) if match_disc else 0
                
                match_ppnbm = re.search(r"PPnBM\s*\(([\d\.,]+)%\)", raw_desc)
                item_data['PPnBM Persen'] = match_ppnbm.group(1) if match_ppnbm else "0,00"
                
                item_data['Harga Total Item'] = parse_currency(raw_total_str)
                data_list.append(item_data)
            
    return data_list

def main():
    output_filename = input("Masukkan nama file output (contoh: hasil.xlsx): ")
    if not output_filename.endswith('.xlsx'):
        output_filename += '.xlsx'
        
    master_data = []
    files = [f for f in os.listdir('.') if f.lower().endswith('.pdf')]
    
    print(f"Ditemukan {len(files)} file PDF. Memulai ekstraksi...")
    
    for f in files:
        try:
            print(f"Memproses: {f}")
            extracted = extract_faktur_data(f)
            master_data.extend(extracted)
        except Exception as e:
            print(f"Gagal memproses {f}: {e}")
            
    if master_data:
        df = pd.DataFrame(master_data)
        
        columns_order = [
            'Nomor Faktur', 'Referensi', 'Tanggal', 'Lokasi',
            'PKP Nama', 'PKP NPWP', 'PKP Alamat',
            'Pembeli Nama', 'Pembeli NPWP', 'Pembeli NIK', 'Pembeli Alamat',
            'Nomor Paspor', 'Identitas Lain', 'Email Pembeli',
            'Kode Barang', 'Nama Barang', 'Qty', 'Satuan', 'Harga Satuan', 
            'Potongan Harga Item', 'PPnBM Persen', 'Harga Total Item',
            'Total Harga Jual', 'Total Potongan', 'DPP', 'PPN', 'PPnBM Total', 'Penandatangan'
        ]
        
        final_cols = [c for c in columns_order if c in df.columns]
        df = df[final_cols]
        
        df.to_excel(output_filename, index=False)
        auto_adjust_excel_width(output_filename)
        print(f"Selesai! Data tersimpan di {output_filename}")
    else:
        print("Tidak ada data yang berhasil diekstrak.")

if __name__ == "__main__":
    main()